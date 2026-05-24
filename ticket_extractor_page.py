#!/usr/bin/env python
# coding: utf-8
"""
ticket_extractor_page.py  —  4G Workbench  |  Ticket Extractor Panel
Fix in this revision:
  - Entry form is no longer squeezed; right panel uses a PanedWindow so
    the form occupies the top ~40% and the tree the bottom ~60% when open.
  - Form slides in/out of a dedicated top pane — always fully visible.
  - All other behaviour unchanged from previous revision.
"""
import csv
import os
import tkinter as tk
from datetime import datetime, date
from tkinter import filedialog, messagebox, ttk
from calendar import monthcalendar, month_abbr
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Palette ───────────────────────────────────────────────────────────────────
BG          = "#FAF3E6"
ACCENT      = "#c9a66b"
ACCENT2     = "#d4b896"
TEXT        = "#3a2f24"
SUBTEXT     = "#7A614A"
WHITE       = "#ffffff"
DONE_BG     = "#c8f2c2"
PANEL_BG    = "#f5ead6"
BORDER      = "#d4b896"
HDR_L       = "#e8f5e9"
HDR_R       = "#fff3e0"
FG_L        = "#2e7d32"
FG_R        = "#ef6c00"
FONT_TITLE  = ("Segoe UI", 11, "bold")
FONT_BOLD   = ("Segoe UI", 10, "bold")
FONT_BODY   = ("Segoe UI", 10)
FONT_SM     = ("Segoe UI", 8)
FONT_SM_B   = ("Segoe UI", 9, "bold")

FRESHSERVICE_BASE = (
    "https://4gcapital.freshservice.com/a/tickets/{tid}?current_tab=details"
)
_TICKET_KEYS  = ["Ticket Id", "ticket_id", "TicketId", "Ticket ID", "ID"]
_SUBJECT_KEYS = ["Subject", "subject"]
_DESC_KEYS    = ["Description", "description", "Desc", "desc"]
_CREATED_KEYS = ["Created Time", "created_time", "Created", "Date"]

def _detect(headers: list, candidates: list) -> str:
    for c in candidates:
        if c in headers:
            return c
    return headers[0] if headers else ""

# ─────────────────────────────────────────────────────────────────────────────
#  Apply consistent ttk dropdown styling
# ─────────────────────────────────────────────────────────────────────────────
def _apply_ttk_styles(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "Warm.TCombobox",
        fieldbackground=WHITE,
        background=ACCENT,
        foreground=TEXT,
        selectbackground=ACCENT,
        selectforeground=WHITE,
        bordercolor=BORDER,
        arrowcolor=TEXT,
        relief="flat",
        padding=(6, 4),
    )
    style.map(
        "Warm.TCombobox",
        fieldbackground=[("readonly", WHITE)],
        foreground=[("readonly", TEXT)],
        background=[("active", ACCENT2)],
        bordercolor=[("focus", ACCENT)],
    )
    style.configure(
        "TScrollbar",
        background=ACCENT2,
        troughcolor=PANEL_BG,
        bordercolor=BORDER,
        arrowcolor=TEXT,
        relief="flat",
    )
    style.configure(
        "Treeview",
        background=WHITE,
        fieldbackground=WHITE,
        foreground=TEXT,
        rowheight=24,
        bordercolor=BORDER,
        relief="flat",
    )
    style.configure(
        "Treeview.Heading",
        background=PANEL_BG,
        foreground=TEXT,
        font=FONT_SM_B,
        relief="flat",
    )
    style.map(
        "Treeview",
        background=[("selected", ACCENT)],
        foreground=[("selected", WHITE)],
    )
    style.map(
        "Treeview.Heading",
        background=[("active", ACCENT2)],
    )
    # Sash styling for PanedWindow
    style.configure("TPanedwindow", background=BORDER)

# ─────────────────────────────────────────────────────────────────────────────
#  Volatile in-memory session DB
# ─────────────────────────────────────────────────────────────────────────────
class SessionDB:
    def __init__(self):
        self.dist_rows:        list[dict] = []
        self.dist_headers:     list[str]  = []
        self.dist_col_ticket   = ""
        self.dist_col_subject  = ""
        self.dist_col_desc     = ""
        self.dist_distributed: list[dict] = []
        self.dist_agents:      list[str]  = []
        self.dist_filename:    str        = ""
        self.dist_done_ids:    set        = set()
        self.ext_rows:        list[dict] = []
        self.ext_headers:     list[str]  = []
        self.ext_col_ticket   = ""
        self.ext_col_subject  = ""
        self.ext_col_desc     = ""
        self.ext_filename:    str        = ""
        self.ext_compilation: list[dict] = []
        self.ext_mode:        str        = "Payment Reconciliation"
        self.ext_left_done_ids: set = set()

DB = SessionDB()

# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _make_tree(parent) -> ttk.Treeview:
    wrap = tk.Frame(parent, bg=WHITE)
    wrap.pack(fill="both", expand=True)
    tree = ttk.Treeview(wrap, show="headings")
    vsb  = ttk.Scrollbar(wrap, orient="vertical",   command=tree.yview)
    hsb  = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right",  fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(side="left",  fill="both", expand=True)
    return tree

def _btn(parent, text, cmd, bg=ACCENT, fg=WHITE, **kw) -> tk.Button:
    kw.setdefault("padx", 8)
    kw.setdefault("pady", 2)
    return tk.Button(
        parent, text=text, command=cmd,
        bg=bg, fg=fg, font=FONT_SM_B, relief="flat",
        cursor="hand2", activebackground=ACCENT2, activeforeground=TEXT,
        **kw)

def _hsep(parent):
    ttk.Separator(parent, orient="horizontal").pack(fill="x", padx=8, pady=3)

def _styled_combobox(parent, textvariable, values, **kw) -> ttk.Combobox:
    kw.setdefault("width", 22)
    kw.setdefault("font", FONT_BODY)
    return ttk.Combobox(
        parent,
        textvariable=textvariable,
        values=values,
        style="Warm.TCombobox",
        state="readonly",
        **kw,
    )

# ─────────────────────────────────────────────────────────────────────────────
#  Shared compact instructions + purpose strip
# ─────────────────────────────────────────────────────────────────────────────
def _make_instructions_strip(parent, mode: str = "extraction") -> tk.Frame:
    outer = tk.Frame(parent, bg="#fff8e1",
                     highlightbackground="#f0c040", highlightthickness=1)
    if mode == "extraction":
        icon    = "📋"
        title   = "Ticket Extraction"
        purpose = (
            "Compile payment details and failed loan records from Freshservice tickets. "
            "Essential for Flexipay, Airtel and any other 3rd party to notify payments and disbursement failures — "
            "outputs a clean Excel report ready for processing."
        )
    else:
        icon    = "🎫"
        title   = "Ticket Distributor"
        purpose = (
            "Evenly split a Freshservice ticket queue across your support agents. "
            "Eliminates manual cherry-picking, ensures fair workload distribution, "
            "and exports a hyperlinked assignment sheet in one click."
        )
    purpose_row = tk.Frame(outer, bg="#fff8e1")
    purpose_row.pack(fill="x", padx=8, pady=(5, 1))
    tk.Label(purpose_row, text=f"{icon}  {title}  — ",
             font=FONT_SM_B, bg="#fff8e1", fg="#6b4c00").pack(side="left")
    tk.Label(purpose_row, text=purpose,
             font=FONT_SM, bg="#fff8e1", fg="#6b4c00", anchor="w",
             wraplength=900, justify="left").pack(side="left", fill="x", expand=True)
    hints_row = tk.Frame(outer, bg="#fff8e1")
    hints_row.pack(fill="x", padx=8, pady=(0, 5))
    tk.Label(hints_row,
             text=("ℹ️  Export CSV from Freshservice with: Ticket ID · Subject · Description · Created Time  "
                   "·  Double-click rows to view description  ·  Right-click rows for actions  "),
             font=FONT_SM, bg="#fff8e1", fg="#6b4c00", anchor="w").pack(anchor="w")
    return outer

# ─────────────────────────────────────────────────────────────────────────────
#  Inline date picker (no Toplevel)
# ─────────────────────────────────────────────────────────────────────────────
class DatePicker(tk.Frame):
    """
    Calendar widget.  Must be a child of the InlineForm frame (NOT the
    scrollable canvas body) so that it renders outside the clipped canvas
    region and is fully visible when shown.
    show()/hide() use pack on the InlineForm itself, inserted before the
    canvas outer frame so it floats in the correct vertical position.
    """
    def __init__(self, parent, textvariable: tk.StringVar, **kw):
        super().__init__(parent, bg=PANEL_BG,
                         highlightbackground=BORDER, highlightthickness=1, **kw)
        self._var = textvariable
        self._today = date.today()
        self._viewing = date(self._today.year, self._today.month, 1)
        self._build()

    def _build(self):
        nav = tk.Frame(self, bg=ACCENT)
        nav.pack(fill="x")
        _btn(nav, "◀", self._prev_month, bg=ACCENT, pady=1).pack(side="left")
        self._nav_lbl = tk.Label(nav, text="", font=FONT_SM_B, bg=ACCENT, fg=WHITE)
        self._nav_lbl.pack(side="left", expand=True)
        _btn(nav, "▶", self._next_month, bg=ACCENT, pady=1).pack(side="right")
        self._cal_frame = tk.Frame(self, bg=PANEL_BG)
        self._cal_frame.pack(padx=4, pady=4)
        for i, d in enumerate(["Mo","Tu","We","Th","Fr","Sa","Su"]):
            tk.Label(self._cal_frame, text=d, font=FONT_SM_B,
                     bg=PANEL_BG, fg=SUBTEXT, width=3).grid(row=0, column=i)
        self._day_btns: list[tk.Button] = []
        for r in range(6):
            for c in range(7):
                b = tk.Button(self._cal_frame, text="", width=3,
                              font=FONT_SM, relief="flat",
                              bg=PANEL_BG, fg=TEXT,
                              activebackground=ACCENT, activeforeground=WHITE,
                              cursor="hand2")
                b.grid(row=r+1, column=c, padx=1, pady=1)
                self._day_btns.append(b)
        btn_row = tk.Frame(self, bg=PANEL_BG)
        btn_row.pack(pady=(0, 4))
        _btn(btn_row, "Today", self._pick_today, pady=1).pack(side="left", padx=4)
        _btn(btn_row, "Close", self.hide, bg=ACCENT2, fg=TEXT, pady=1).pack(side="left", padx=4)
        self._refresh()

    def _refresh(self):
        v = self._viewing
        self._nav_lbl.config(text=f"{month_abbr[v.month]} {v.year}")
        weeks = monthcalendar(v.year, v.month)
        while len(weeks) < 6:
            weeks.append([0]*7)
        for idx, btn in enumerate(self._day_btns):
            r, c = divmod(idx, 7)
            day = weeks[r][c] if r < len(weeks) else 0
            if day == 0:
                btn.config(text="", state="disabled", bg=PANEL_BG, fg=TEXT)
            else:
                d = date(v.year, v.month, day)
                is_today    = (d == self._today)
                is_selected = (self._var.get() == d.strftime("%Y-%m-%d"))
                bg = ACCENT  if is_selected else \
                     ACCENT2 if is_today    else PANEL_BG
                fg = WHITE if (is_selected or is_today) else TEXT
                btn.config(text=str(day), state="normal", bg=bg, fg=fg,
                           command=lambda dy=day: self._pick(dy))

    def _pick(self, day: int):
        v = self._viewing
        chosen = date(v.year, v.month, day)
        self._var.set(chosen.strftime("%Y-%m-%d"))
        self._refresh()
        self.hide()

    def _pick_today(self):
        self._var.set(self._today.strftime("%Y-%m-%d"))
        self._viewing = date(self._today.year, self._today.month, 1)
        self._refresh()
        self.hide()

    def _prev_month(self):
        y, m = self._viewing.year, self._viewing.month
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        self._viewing = date(y, m, 1)
        self._refresh()

    def _next_month(self):
        y, m = self._viewing.year, self._viewing.month
        m += 1
        if m == 13:
            m, y = 1, y + 1
        self._viewing = date(y, m, 1)
        self._refresh()

    def show(self):
        self.pack(fill="x", padx=8, pady=(0, 6))

    def hide(self):
        self.pack_forget()

# ─────────────────────────────────────────────────────────────────────────────
#  Inline description viewer
# ─────────────────────────────────────────────────────────────────────────────
class DescriptionPane(tk.Frame):
    def __init__(self, parent, on_hide=None, **kw):
        super().__init__(parent, bg=PANEL_BG,
                         highlightbackground=BORDER, highlightthickness=1, **kw)
        self._on_hide = on_hide
        hdr = tk.Frame(self, bg=HDR_L)
        hdr.pack(fill="x")
        self._title_lbl = tk.Label(hdr, text="Description",
                                   font=FONT_SM_B, bg=HDR_L, fg=FG_L, anchor="w")
        self._title_lbl.pack(side="left", padx=8, pady=3)
        self._copy_btn = _btn(hdr, "✂ Copy", self._copy, bg=ACCENT)
        self._copy_btn.pack(side="right", padx=4, pady=2)
        _btn(hdr, "✕ Close", self.hide, bg=ACCENT2, fg=TEXT).pack(
            side="right", padx=2, pady=2)
        txt_wrap = tk.Frame(self, bg=PANEL_BG)
        txt_wrap.pack(fill="both", expand=True, padx=4, pady=4)
        vsb = ttk.Scrollbar(txt_wrap)
        vsb.pack(side="right", fill="y")
        self._txt = tk.Text(
            txt_wrap, font=FONT_BODY, bg=WHITE, fg=TEXT, relief="flat",
            wrap="word", height=5, bd=1,
            highlightbackground=BORDER, highlightthickness=1,
            yscrollcommand=vsb.set)
        self._txt.pack(fill="both", expand=True)
        vsb.config(command=self._txt.yview)
        self._content = ""

    def show(self, ticket_id: str, subject: str, description: str):
        self._title_lbl.config(text=f"#{ticket_id}  ·  {subject}")
        self._content = description or "(No description)"
        self._txt.config(state="normal")
        self._txt.delete("1.0", "end")
        self._txt.insert("1.0", self._content)
        # Visibility is driven by ExtractionPanel._show_desc_pane()

    def hide(self):
        if self._on_hide:
            self._on_hide()

    def _copy(self):
        self.clipboard_clear()
        self.clipboard_append(self._content)
        self._copy_btn.config(text="✔ Copied!")
        self.after(1500, lambda: self._copy_btn.config(text="✂ Copy"))

# ─────────────────────────────────────────────────────────────────────────────
#  FormHost — a dedicated container that sits ABOVE the right tree.
#  When a form is shown, this frame becomes visible and sized properly.
#  The PanedWindow in ExtractionPanel ensures it gets real estate.
# ─────────────────────────────────────────────────────────────────────────────
class FormHost(tk.Frame):
    """
    Holds exactly one inline form at a time.
    Visibility is managed externally by ExtractionPanel via show/hide.
    The frame is always packed inside the top pane of a PanedWindow,
    so it always has room — no fighting with the tree.
    """
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=PANEL_BG, **kw)
        self._current: tk.Frame | None = None

    def attach(self, form: tk.Frame):
        """Pack form inside this host, removing the previous one first."""
        if self._current and self._current is not form:
            self._current.pack_forget()
        form.pack(fill="both", expand=True)
        self._current = form

    def detach(self):
        if self._current:
            self._current.pack_forget()
            self._current = None

# ─────────────────────────────────────────────────────────────────────────────
#  InlineForm base — renders inside FormHost (no expand fighting with tree)
# ─────────────────────────────────────────────────────────────────────────────
class InlineForm(tk.Frame):
    def __init__(self, parent_host: FormHost, title: str, on_hide=None, **kw):
        super().__init__(parent_host, bg=PANEL_BG, **kw)
        self._host    = parent_host
        self._on_hide = on_hide

        # Header row — title on left, action buttons + close on right
        self._hdr = tk.Frame(self, bg=HDR_R)
        self._hdr.pack(fill="x")
        tk.Label(self._hdr, text=title, font=FONT_SM_B,
                 bg=HDR_R, fg=FG_R, anchor="w").pack(side="left", padx=10, pady=4)
        # Close button — subclass buttons pack to the right BEFORE this is called
        # so they appear left of Close. We store it and pack last.
        self._close_btn = _btn(self._hdr, "✕ Close", self.hide,
                               bg=ACCENT2, fg=TEXT, padx=8, pady=3)
        # (packed after subclass __init__ adds its buttons via _add_header_btn)

        # Scrollable body
        _outer = tk.Frame(self, bg=PANEL_BG)
        _outer.pack(fill="both", expand=True)
        self._canvas = tk.Canvas(_outer, bg=PANEL_BG, highlightthickness=0)
        _vsb = ttk.Scrollbar(_outer, orient="vertical", command=self._canvas.yview)
        _vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._canvas.configure(yscrollcommand=_vsb.set)
        self._body = tk.Frame(self._canvas, bg=PANEL_BG)
        self._body_id = self._canvas.create_window((0, 0), window=self._body, anchor="nw")
        self._body.bind("<Configure>", self._on_body_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)

    def _pack_close(self):
        """Call at end of subclass __init__ after injecting any header buttons."""
        self._close_btn.pack(side="right", padx=6, pady=4)

    def _add_header_btn(self, text, cmd, bg=ACCENT, fg=WHITE):
        """Add a button to the right side of the header (before Close)."""
        b = _btn(self._hdr, text, cmd, bg=bg, fg=fg, padx=8, pady=3)
        b.pack(side="right", padx=(0, 4), pady=4)
        return b

    def _on_body_configure(self, _=None):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self._canvas.itemconfig(self._body_id, width=event.width)

    def show(self):
        self._host.attach(self)

    def hide(self):
        self._host.detach()
        if self._on_hide:
            self._on_hide()

    def _field(self, row: int, label: str, var: tk.StringVar,
               readonly=False, combo=None, col=0):
        tk.Label(self._body, text=label, font=FONT_SM_B, bg=PANEL_BG,
                 fg=TEXT, anchor="w").grid(
            row=row, column=col * 2, sticky="w", padx=(10, 6), pady=10)
        if combo:
            w = _styled_combobox(self._body, var, combo, width=24)
        else:
            w = tk.Entry(
                self._body, textvariable=var, font=FONT_BODY, width=26,
                state="disabled" if readonly else "normal",
                bg="#f0e8d8" if readonly else WHITE, fg=TEXT,
                relief="flat", bd=0,
                highlightbackground=BORDER, highlightthickness=1,
                disabledbackground="#f0e8d8", disabledforeground=SUBTEXT)
        w.grid(row=row, column=col * 2 + 1, sticky="ew", padx=(0, 20), pady=10)
        return var

# ─────────────────────────────────────────────────────────────────────────────
#  Payment Reconciliation form
# ─────────────────────────────────────────────────────────────────────────────
class ReconForm(InlineForm):
    def __init__(self, parent_host: FormHost, on_save, **kw):
        super().__init__(parent_host, "➕  Payment Reconciliation Entry", **kw)
        self._on_save = on_save
        self._v: dict[str, tk.StringVar] = {}

        # ── Header buttons (right-to-left pack order → left-to-right visually)
        # Close is packed last by _pack_close(), so add Save first so it sits left of Close
        self._add_header_btn("💾 Save", self._save, bg=ACCENT,  fg=WHITE)
        self._pack_close()

        b = self._body
        b.columnconfigure(1, weight=1)
        b.columnconfigure(3, weight=1)

        self._v["ticket_id"] = tk.StringVar()
        self._v["amount"]    = tk.StringVar()
        self._v["amount"].trace_add("write", self._strip_amount_commas)
        self._stripping = False

        # Row 0 — Ticket # | Amount
        self._field(0, "Ticket #",     self._v["ticket_id"], readonly=True, col=0)
        self._field(0, "Amount (UGX)", self._v["amount"],    col=1)

        # Row 1 — Phone | Payment Date (pre-filled, no picker needed)
        self._v["phone"] = tk.StringVar(value="+256")
        self._field(1, "Number Used to Pay", self._v["phone"], col=0)

        self._v["payment_date"] = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        self._field(1, "Payment Date", self._v["payment_date"], col=1)

        # Row 2 — Payment Mode | Transaction ID
        self._v["payment_mode"] = tk.StringVar(value="Flexipay")
        self._field(2, "Payment Mode", self._v["payment_mode"],
                    combo=["Flexipay", "Beyonic", "Airtel", "Bank"], col=0)
        # Transaction ID — Row 2, right column
        self._v["transaction_id"] = tk.StringVar()
        self._field(2, "Transaction ID", self._v["transaction_id"], col=1)

    def _strip_amount_commas(self, *_):
        if self._stripping:
            return
        raw = self._v["amount"].get()
        if "," in raw:
            self._stripping = True
            self._v["amount"].set(raw.replace(",", ""))
            self._stripping = False

    def prime(self, ticket_id: str):
        self._v["ticket_id"].set(ticket_id)
        self._v["amount"].set("")
        self._v["phone"].set("+256")
        self._v["payment_date"].set(date.today().strftime("%Y-%m-%d"))
        self._v["payment_mode"].set("Flexipay")
        self._v["transaction_id"].set("")

    def _save(self):
        d = {k: v.get().strip() for k, v in self._v.items()}
        for fld, lbl in [("amount",       "Amount"),
                         ("phone",        "Number Used to Pay"),
                         ("payment_date", "Payment Date")]:
            if not d.get(fld):
                messagebox.showwarning("Missing", f"{lbl} is required.")
                return
        self._on_save(d)
        self.hide()

# ─────────────────────────────────────────────────────────────────────────────
#  Failed Loans form
# ─────────────────────────────────────────────────────────────────────────────
class FailedForm(InlineForm):
    def __init__(self, parent_host: FormHost, on_save, **kw):
        super().__init__(parent_host, "➕  Failed Loan Entry", **kw)
        self._on_save = on_save
        self._v: dict[str, tk.StringVar] = {}

        # Header buttons
        self._add_header_btn("💾 Save", self._save, bg=ACCENT, fg=WHITE)
        self._pack_close()

        b = self._body
        b.columnconfigure(1, weight=1)
        self._v["ticket_id"] = tk.StringVar()
        self._v["phone"]     = tk.StringVar(value="+256")
        self._field(0, "Ticket #",     self._v["ticket_id"], readonly=True)
        self._field(1, "Phone Number", self._v["phone"])

    def prime(self, ticket_id: str):
        self._v["ticket_id"].set(ticket_id)
        self._v["phone"].set("+256")

    def _save(self):
        d = {k: v.get().strip() for k, v in self._v.items()}
        if not d.get("phone"):
            messagebox.showwarning("Missing", "Phone number is required.")
            return
        self._on_save(d)
        self.hide()

# ─────────────────────────────────────────────────────────────────────────────
#  Pill-style radio button group
# ─────────────────────────────────────────────────────────────────────────────
class PillRadio(tk.Frame):
    def __init__(self, parent, options: list[str],
                 variable: tk.StringVar, command=None, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._var  = variable
        self._cmd  = command
        self._btns: dict[str, tk.Button] = {}
        for opt in options:
            b = tk.Button(
                self, text=opt, font=FONT_SM_B, relief="flat",
                cursor="hand2", padx=12, pady=4,
                command=lambda o=opt: self._select(o))
            b.pack(side="left", padx=1)
            self._btns[opt] = b
        self._refresh()

    def _select(self, opt: str):
        self._var.set(opt)
        self._refresh()
        if self._cmd:
            self._cmd()

    def _refresh(self):
        for opt, btn in self._btns.items():
            if self._var.get() == opt:
                btn.config(bg=ACCENT, fg=WHITE,
                           highlightbackground=ACCENT, highlightthickness=2)
            else:
                btn.config(bg=WHITE, fg=TEXT,
                           highlightbackground=BORDER, highlightthickness=1)

# ─────────────────────────────────────────────────────────────────────────────
#  Agent Manager Widget
# ─────────────────────────────────────────────────────────────────────────────
class AgentManager(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=PANEL_BG,
                         highlightbackground=BORDER, highlightthickness=1, **kw)
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=HDR_L)
        hdr.pack(fill="x")
        tk.Label(hdr, text="👥  Agent List", font=FONT_SM_B,
                 bg=HDR_L, fg=FG_L).pack(side="left", padx=8, pady=4)
        self._count_lbl = tk.Label(hdr, text="0 agents",
                                   font=FONT_SM, bg=HDR_L, fg=SUBTEXT)
        self._count_lbl.pack(side="right", padx=8)

        input_row = tk.Frame(self, bg=PANEL_BG)
        input_row.pack(fill="x", padx=8, pady=(6, 4))
        self._name_var = tk.StringVar()
        self._entry = tk.Entry(
            input_row, textvariable=self._name_var,
            font=FONT_BODY, bg=WHITE, fg=TEXT,
            relief="flat", bd=0,
            highlightbackground=BORDER, highlightthickness=1,
            width=12)
        self._entry.pack(side="left", ipady=4)
        self._entry.bind("<Return>", lambda _: self._add_agent())
        _btn(input_row, "➕ Add", self._add_agent,
             padx=10, pady=4).pack(side="left", padx=(6, 0))
        _btn(input_row, "🗑 Clear All", self._clear_all,
             bg=ACCENT2, fg=TEXT, padx=8, pady=4).pack(side="left", padx=(4, 0))

        list_frame = tk.Frame(self, bg=WHITE,
                              highlightbackground=BORDER, highlightthickness=1)
        list_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        vsb = ttk.Scrollbar(list_frame)
        vsb.pack(side="right", fill="y")
        self._listbox = tk.Listbox(
            list_frame, font=FONT_BODY, bg=WHITE, fg=TEXT,
            selectbackground=ACCENT, selectforeground=WHITE,
            relief="flat", bd=0, highlightthickness=0,
            activestyle="none", height=5, yscrollcommand=vsb.set)
        self._listbox.pack(side="left", fill="both", expand=True)
        vsb.config(command=self._listbox.yview)
        self._listbox.bind("<Double-1>", lambda _: self._remove_selected())

        tk.Label(self, text="Double-click an agent to remove  ·  Press Enter to add",
                 font=FONT_SM, bg=PANEL_BG, fg=SUBTEXT).pack(pady=(0, 4))

    def _add_agent(self):
        name = self._name_var.get().strip()
        if not name:
            return
        if name in list(self._listbox.get(0, "end")):
            messagebox.showwarning("Duplicate", f'"{name}" is already in the list.')
            return
        self._listbox.insert("end", name)
        DB.dist_agents.append(name)
        self._name_var.set("")
        self._entry.focus()
        self._update_count()

    def _remove_selected(self):
        sel = self._listbox.curselection()
        if not sel:
            return
        idx  = sel[0]
        name = self._listbox.get(idx)
        self._listbox.delete(idx)
        if name in DB.dist_agents:
            DB.dist_agents.remove(name)
        self._update_count()

    def _clear_all(self):
        if not self._listbox.size():
            return
        if messagebox.askyesno("Clear Agents", "Remove all saved agents?"):
            self._listbox.delete(0, "end")
            DB.dist_agents.clear()
            self._update_count()

    def _update_count(self):
        n = self._listbox.size()
        self._count_lbl.config(text=f"{n} agent{'s' if n != 1 else ''}")

    def get_agents(self) -> list[str]:
        return list(self._listbox.get(0, "end"))

    def restore(self, agents: list[str]):
        self._listbox.delete(0, "end")
        for a in agents:
            self._listbox.insert("end", a)
        self._update_count()

# ─────────────────────────────────────────────────────────────────────────────
#  TICKET DISTRIBUTOR
# ─────────────────────────────────────────────────────────────────────────────
class TicketsDistributionPanel(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._build()

    def _build(self):
        _make_instructions_strip(self, mode="distributor").pack(
            fill="x", padx=6, pady=(4, 2))

        toolbar = tk.Frame(self, bg="#e8d5b0",
                           highlightbackground=BORDER, highlightthickness=1)
        toolbar.pack(fill="x", padx=0, pady=(0, 2))
        tk.Label(toolbar, text="🎫  Ticket Distributor",
                 font=FONT_BOLD, bg="#e8d5b0", fg=TEXT).pack(
                     side="left", padx=10, pady=5)
        self._btn_export = _btn(toolbar, "📤 Export CSV", self._export_csv,
                                bg=ACCENT2, fg=TEXT)
        self._btn_export.config(state="disabled")
        self._btn_export.pack(side="right", padx=4, pady=4)
        _btn(toolbar, "👥 Distribute", self._distribute).pack(side="right", padx=4, pady=4)
        _btn(toolbar, "🔄 Reset", self._reset, bg=ACCENT2, fg=TEXT).pack(side="right", padx=4, pady=4)

        setup = tk.Frame(self, bg=PANEL_BG,
                         highlightbackground=BORDER, highlightthickness=1)
        setup.pack(fill="x", padx=6, pady=4)
        row_top = tk.Frame(setup, bg=PANEL_BG)
        row_top.pack(fill="x", padx=6, pady=(4, 2))
        _btn(row_top, "📂 Upload CSV", self._load_csv).pack(side="left")
        self._lbl_file = tk.Label(row_top, text="No file loaded",
                                  font=FONT_SM, bg=PANEL_BG, fg=SUBTEXT)
        self._lbl_file.pack(side="left", padx=8)
        self._lbl_status = tk.Label(row_top, text="",
                                    font=FONT_SM, bg=PANEL_BG, fg=FG_L)
        self._lbl_status.pack(side="left")

        self._col_frame = tk.Frame(setup, bg=PANEL_BG)
        self._col_vars  = {k: tk.StringVar() for k in ("ticket", "subject", "desc")}
        col_labels      = {"ticket": "Ticket ID:", "subject": "Subject:", "desc": "Description:"}
        self._col_cbs: dict[str, ttk.Combobox] = {}
        for i, (key, lbl) in enumerate(col_labels.items()):
            tk.Label(self._col_frame, text=lbl, font=FONT_SM,
                     bg=PANEL_BG, fg=TEXT).grid(
                         row=0, column=i * 2, sticky="w", padx=(8 if i else 0, 4))
            cb = _styled_combobox(self._col_frame, self._col_vars[key], [], width=16)
            cb.bind("<<ComboboxSelected>>", lambda _: self._populate_tree())
            cb.grid(row=0, column=i * 2 + 1, padx=(0, 12), pady=2)
            self._col_cbs[key] = cb

        self._agent_mgr = AgentManager(setup)
        self._agent_mgr.pack(fill="x", padx=6, pady=(4, 6))

        self._lbl_dist = tk.Label(self, text="", font=FONT_SM, bg=BG, fg=FG_L)
        self._lbl_dist.pack(anchor="w", padx=8)

        tree_frame = tk.Frame(self, bg=BG,
                              highlightbackground=BORDER, highlightthickness=1)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=(0, 4))
        t_hdr = tk.Frame(tree_frame, bg=HDR_L)
        t_hdr.pack(fill="x")
        tk.Label(t_hdr, text="Loaded / Distributed Tickets",
                 font=FONT_SM_B, bg=HDR_L, fg=FG_L).pack(side="left", padx=6, pady=3)
        tk.Label(t_hdr, text="🟢 Mark as Done is independent — right-click any row",
                 font=FONT_SM, bg=HDR_L, fg=FG_L).pack(side="right", padx=8, pady=3)
        self._tree = _make_tree(tree_frame)
        _cols = ("#", "Agent", "Ticket ID", "Subject")
        self._tree["columns"] = _cols
        for c, w in zip(_cols, [44, 130, 100, 400]):
            self._tree.heading(c, text=c)
            self._tree.column(c, width=w, minwidth=30, anchor="w")
        self._tree.tag_configure("alt",  background="#fdf8f0")
        self._tree.tag_configure("done", background=DONE_BG)
        self._tree.bind("<Button-3>", self._on_tree_rclick)

    def show(self):
        if DB.dist_rows:
            self._populate_tree()
            self._lbl_file.config(text=DB.dist_filename or "file loaded", fg=FG_L)
            self._lbl_status.config(text=f"✅  {len(DB.dist_rows)} tickets loaded")
        if DB.dist_agents:
            self._agent_mgr.restore(DB.dist_agents)
        if DB.dist_distributed:
            self._btn_export.config(state="normal")
            self._refresh_dist_tree()

    def _load_csv(self):
        path = filedialog.askopenfilename(
            title="Select Freshservice CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader          = csv.DictReader(f)
                DB.dist_headers = list(reader.fieldnames or [])
                DB.dist_rows    = list(reader)
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            return
        if not DB.dist_rows:
            messagebox.showwarning("Empty", "The file appears to be empty.")
            return
        DB.dist_col_ticket  = _detect(DB.dist_headers, _TICKET_KEYS)
        DB.dist_col_subject = _detect(DB.dist_headers, _SUBJECT_KEYS)
        DB.dist_col_desc    = _detect(DB.dist_headers, _DESC_KEYS)
        DB.dist_filename    = os.path.basename(path)
        DB.dist_done_ids    = set()
        ambiguous = (
            DB.dist_col_ticket  not in _TICKET_KEYS or
            DB.dist_col_subject not in _SUBJECT_KEYS or
            DB.dist_col_desc    not in _DESC_KEYS
        )
        if ambiguous:
            for key, cb in self._col_cbs.items():
                cb["values"] = DB.dist_headers
            self._col_vars["ticket"].set(DB.dist_col_ticket)
            self._col_vars["subject"].set(DB.dist_col_subject)
            self._col_vars["desc"].set(DB.dist_col_desc)
            self._col_frame.pack(fill="x", padx=6, pady=(0, 4))
        else:
            self._col_frame.pack_forget()
        self._populate_tree()
        self._lbl_file.config(text=DB.dist_filename, fg=FG_L)
        self._lbl_status.config(text=f"✅  {len(DB.dist_rows)} tickets loaded")
        DB.dist_distributed = []
        self._lbl_dist.config(text="")
        self._btn_export.config(state="disabled")

    def _resolve_cols(self):
        if self._col_frame.winfo_ismapped():
            return (self._col_vars["ticket"].get(),
                    self._col_vars["subject"].get(),
                    self._col_vars["desc"].get())
        return DB.dist_col_ticket, DB.dist_col_subject, DB.dist_col_desc

    def _populate_tree(self):
        self._tree.delete(*self._tree.get_children())
        tc, sc, _ = self._resolve_cols()
        for i, row in enumerate(DB.dist_rows, 1):
            tag = "alt" if i % 2 == 0 else ""
            tid = row.get(tc, "")
            if tid in DB.dist_done_ids:
                tag = "done"
            self._tree.insert("", "end", iid=str(i),
                              values=(i, "", tid, row.get(sc, "")), tags=(tag,))

    def _on_tree_rclick(self, event):
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        self._tree.selection_set(iid)
        vals = self._tree.item(iid, "values")
        tid  = vals[2] if len(vals) > 2 else ""
        menu = tk.Menu(self, tearoff=0, bg=BG, fg=TEXT, font=FONT_BODY,
                       activebackground=ACCENT, activeforeground=WHITE)
        if tid in DB.dist_done_ids:
            menu.add_command(label="↩  Unmark Done",
                             command=lambda t=tid, i=iid: self._set_done(t, i, False))
        else:
            menu.add_command(label="✅  Mark as Done  🟢",
                             command=lambda t=tid, i=iid: self._set_done(t, i, True))
        menu.post(event.x_root, event.y_root)

    def _set_done(self, tid: str, iid: str, done: bool):
        if done:
            DB.dist_done_ids.add(tid)
            self._tree.item(iid, tags=("done",))
        else:
            DB.dist_done_ids.discard(tid)
            idx = int(iid)
            tag = "alt" if idx % 2 == 0 else ""
            self._tree.item(iid, tags=(tag,))

    def _distribute(self):
        if not DB.dist_rows:
            messagebox.showwarning("No Data", "Load a CSV file first.")
            return
        names = self._agent_mgr.get_agents()
        if not names:
            messagebox.showwarning("No Agents", "Add at least one agent name first.")
            return
        DB.dist_agents = names[:]
        tc, sc, _ = self._resolve_cols()
        n = len(DB.dist_rows)
        k = len(names)
        base, extra = divmod(n, k)
        DB.dist_distributed = []
        serial = 1
        for agent_idx, name in enumerate(names):
            count = base + (1 if agent_idx < extra else 0)
            start = sum(base + (1 if j < extra else 0) for j in range(agent_idx))
            for row_idx in range(start, start + count):
                row = DB.dist_rows[row_idx]
                DB.dist_distributed.append({
                    "serial":    serial,
                    "agent":     name,
                    "ticket_id": row.get(tc, ""),
                    "subject":   row.get(sc, ""),
                })
                serial += 1
        counts  = {}
        for d in DB.dist_distributed:
            counts[d["agent"]] = counts.get(d["agent"], 0) + 1
        summary = "  |  ".join(f"{a}: {c}" for a, c in counts.items())
        self._lbl_dist.config(
            text=f"✔  {len(DB.dist_distributed)} distributed — {summary}")
        self._btn_export.config(state="normal")
        self._refresh_dist_tree()

    def _refresh_dist_tree(self):
        self._tree.delete(*self._tree.get_children())
        for d in DB.dist_distributed:
            i   = d["serial"]
            tid = d["ticket_id"]
            tag = "done" if tid in DB.dist_done_ids else \
                  "alt"  if i % 2 == 0 else ""
            self._tree.insert("", "end", iid=str(i),
                              values=(i, d["agent"], tid, d["subject"]),
                              tags=(tag,))

    def _export_csv(self):
        if not DB.dist_distributed:
            messagebox.showwarning("Nothing to Export", "Run distribute first.")
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"ticket_distribution-{ts}.csv"
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", initialfile=name,
            filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Serial Number", "Agent Name", "Ticket ID", "Done"])
                for d in DB.dist_distributed:
                    link = FRESHSERVICE_BASE.format(tid=d["ticket_id"])
                    done = "Yes" if d["ticket_id"] in DB.dist_done_ids else ""
                    w.writerow([d["serial"], d["agent"],
                                f'=HYPERLINK("{link}","{d["ticket_id"]}")', done])
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _reset(self):
        DB.dist_rows        = []
        DB.dist_distributed = []
        DB.dist_headers     = []
        DB.dist_agents      = []
        DB.dist_filename    = ""
        DB.dist_done_ids    = set()
        self._tree.delete(*self._tree.get_children())
        self._col_frame.pack_forget()
        self._agent_mgr.restore([])
        self._lbl_file.config(text="No file loaded", fg=SUBTEXT)
        self._lbl_status.config(text="")
        self._lbl_dist.config(text="")
        self._btn_export.config(state="disabled")

# ─────────────────────────────────────────────────────────────────────────────
#  TICKET EXTRACTION  — RIGHT PANEL REBUILT WITH PanedWindow
#
#  Layout of right_wrap:
#    ┌──────────────────────────────────┐
#    │  r_hdr  (header + mode dropdown) │
#    │  act_bar (buttons)               │
#    ├──────────────────────────────────┤  ← PanedWindow sash
#    │  FormHost  (form lives here)     │  top pane — hidden when no form
#    ├──────────────────────────────────┤  ← sash
#    │  right tree                      │  bottom pane — always visible
#    └──────────────────────────────────┘
# ─────────────────────────────────────────────────────────────────────────────
class ExtractionPanel(tk.Frame):
    RECON_COLS  = ("#", "Amount (UGX)", "Number Used to Pay",
                   "Payment Date", "Payment Mode", "Transaction ID", "Ticket #")
    FAILED_COLS = ("#", "Phone Number", "Ticket #")

    # Height reserved for the form pane when open (pixels)
    # 250px fits all 3 field rows + buttons; calendar expands below via scroll
    _FORM_PANE_H = 200

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._row_map:  dict[str, dict] = {}
        self._mode_var = tk.StringVar(value=DB.ext_mode)
        self._form_open = False
        self._build()

    def _build(self):
        _make_instructions_strip(self, mode="extraction").pack(
            fill="x", padx=6, pady=(4, 2))

        # ── Toolbar ───────────────────────────────────────────────────────
        toolbar = tk.Frame(self, bg="#e8d5b0",
                           highlightbackground=BORDER, highlightthickness=1)
        toolbar.pack(fill="x", padx=0, pady=(0, 2))
        tk.Label(toolbar, text="📋  Ticket Extraction",
                 font=FONT_BOLD, bg="#e8d5b0", fg=TEXT).pack(
                     side="left", padx=10, pady=5)

        # ── File row ──────────────────────────────────────────────────────
        file_row = tk.Frame(self, bg=PANEL_BG,
                            highlightbackground=BORDER, highlightthickness=1)
        file_row.pack(fill="x", padx=6, pady=(0, 2))
        _btn(file_row, "📂 Upload CSV", self._load_csv).pack(side="left", padx=6, pady=4)
        self._lbl_file = tk.Label(file_row, text="No file loaded",
                                  font=FONT_SM, bg=PANEL_BG, fg=SUBTEXT)
        self._lbl_file.pack(side="left", padx=4, pady=4)

        # ── Two-column results area ───────────────────────────────────────
        results = tk.Frame(self, bg=BG)
        results.pack(fill="both", expand=True, padx=6, pady=(0, 4))
        results.grid_columnconfigure(0, weight=1, uniform="ex")
        results.grid_columnconfigure(1, weight=1, uniform="ex")
        results.grid_rowconfigure(0, weight=1)

        # ─── LEFT panel — PanedWindow mirrors right panel ─────────────────
        left_wrap = tk.Frame(results, bg=BG, bd=1, relief="ridge")
        left_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 3))

        l_hdr = tk.Frame(left_wrap, bg=HDR_L)
        l_hdr.pack(fill="x")
        tk.Label(l_hdr, text="Tickets List", font=FONT_SM_B,
                 bg=HDR_L, fg=FG_L).pack(side="left", padx=6, pady=3)
        self._lbl_left_count = tk.Label(l_hdr, text="Rows: 0",
                                        font=FONT_SM, bg=HDR_L, fg=FG_L)
        self._lbl_left_count.pack(side="right", padx=6)
        tk.Label(l_hdr, text="Dbl-click: description  |  R-click: actions",
                 font=FONT_SM, bg=HDR_L, fg=FG_L).pack(side="right", padx=6)

        # PanedWindow: desc pane (top, collapsed until used) + tree (bottom)
        self._left_paned = tk.PanedWindow(
            left_wrap,
            orient=tk.VERTICAL,
            sashwidth=6,
            sashrelief="flat",
            bg=BORDER,
            handlesize=0,
        )
        self._left_paned.pack(fill="both", expand=True)

        # Top pane — DescriptionPane (collapsed to 0 until a row is opened)
        self._desc_pane = DescriptionPane(self._left_paned,
                                          on_hide=self._hide_desc_pane)
        self._left_paned.add(self._desc_pane, minsize=0, stretch="never")
        self._left_paned.paneconfig(self._desc_pane, height=0)

        # Bottom pane — ticket tree (always visible)
        left_tree_pane = tk.Frame(self._left_paned, bg=WHITE)
        self._left_paned.add(left_tree_pane, minsize=60, stretch="always")

        self._left_tree = _make_tree(left_tree_pane)
        left_cols = ("#", "Ticket ID", "Subject", "Created")
        self._left_tree["columns"] = left_cols
        for c, w in zip(left_cols, [38, 90, 230, 120]):
            self._left_tree.heading(c, text=c)
            self._left_tree.column(c, width=w, minwidth=28, anchor="w")
        self._left_tree.tag_configure("alt",  background="#fdf8f0")
        self._left_tree.tag_configure("done", background=DONE_BG)
        self._left_tree.bind("<Double-1>", self._on_left_double)
        self._left_tree.bind("<Button-3>", self._on_left_rclick)

        # ─── RIGHT panel — PanedWindow layout ────────────────────────────
        right_wrap = tk.Frame(results, bg=BG, bd=1, relief="ridge")
        right_wrap.grid(row=0, column=1, sticky="nsew", padx=(3, 0))

        # Header
        r_hdr = tk.Frame(right_wrap, bg=HDR_R)
        r_hdr.pack(fill="x")
        tk.Label(r_hdr, text="Compilation List", font=FONT_SM_B,
                 bg=HDR_R, fg=FG_R).pack(side="left", padx=6, pady=3)
        self._lbl_right_count = tk.Label(r_hdr, text="Rows: 0",
                                         font=FONT_SM, bg=HDR_R, fg=FG_R)
        self._lbl_right_count.pack(side="right", padx=6)

        # Mode selector dropdown
        mode_wrap = tk.Frame(r_hdr, bg=HDR_R)
        mode_wrap.pack(side="right", padx=(0, 8), pady=3)
        tk.Label(mode_wrap, text="Mode:", font=FONT_SM_B,
                 bg=HDR_R, fg=FG_R).pack(side="left", padx=(0, 4))
        self._mode_cb = _styled_combobox(
            mode_wrap, self._mode_var,
            ["Payment Reconciliation", "Failed Loans"], width=20)
        self._mode_cb.pack(side="left")
        self._mode_cb.bind("<<ComboboxSelected>>", lambda _: self._mode_changed())

        # Action bar
        act_bar = tk.Frame(right_wrap, bg=BG)
        act_bar.pack(fill="x", padx=4, pady=3)
        _btn(act_bar, "➕ Add Entry", self._open_form).pack(side="left")
        self._btn_export = _btn(act_bar, "📥 Export Excel",
                                self._export_excel, bg=ACCENT2, fg=TEXT)
        self._btn_export.config(state="disabled")
        self._btn_export.pack(side="left", padx=4)
        _btn(act_bar, "🗑 Clear", self._clear_compilation,
             bg=ACCENT2, fg=TEXT).pack(side="left")
        tk.Label(act_bar, text="R-click row → Remove / Done",
                 font=FONT_SM, bg=BG, fg=SUBTEXT).pack(side="right", padx=6)

        # ── PanedWindow: form pane (top) + tree pane (bottom) ────────────
        self._paned = tk.PanedWindow(
            right_wrap,
            orient=tk.VERTICAL,
            sashwidth=6,
            sashrelief="flat",
            bg=BORDER,
            handlesize=0,
        )
        self._paned.pack(fill="both", expand=True)

        # Top pane — FormHost (initially hidden by collapsing its pane)
        self._form_host = FormHost(self._paned)
        self._paned.add(self._form_host, minsize=0, stretch="never")
        # Hide initially — set height to 0
        self._paned.paneconfig(self._form_host, height=0)

        # Bottom pane — always-visible right tree
        tree_pane = tk.Frame(self._paned, bg=WHITE)
        self._paned.add(tree_pane, minsize=60, stretch="always")

        self._right_tree = _make_tree(tree_pane)
        self._right_tree.tag_configure("alt",  background="#fdf8f0")
        self._right_tree.tag_configure("done", background=DONE_BG)
        self._right_tree.bind("<Button-3>", self._on_right_rclick)
        self._rebuild_right_cols()

        # Build forms — both children of the FormHost
        self._recon_form  = ReconForm(self._form_host,  self._save_entry,
                                      on_hide=self._on_form_hidden)
        self._failed_form = FailedForm(self._form_host, self._save_entry,
                                       on_hide=self._on_form_hidden)

    # ── Form show/hide with PanedWindow resize ────────────────────────────
    def _show_form_pane(self):
        """Expand the right form pane to _FORM_PANE_H pixels."""
        self._paned.paneconfig(self._form_host, height=self._FORM_PANE_H)
        self._form_open = True

    def _hide_form_pane(self):
        """Collapse the right form pane to 0 height."""
        self._paned.paneconfig(self._form_host, height=0)
        self._form_open = False

    def _on_form_hidden(self):
        self._hide_form_pane()

    # ── Desc pane show/hide (left panel) — mirrors right panel ───────────
    _DESC_PANE_H = 110   # pixels — half the original to save vertical space

    def _show_desc_pane(self):
        self._left_paned.paneconfig(self._desc_pane, height=self._DESC_PANE_H)

    def _hide_desc_pane(self):
        self._left_paned.paneconfig(self._desc_pane, height=0)

    # ── Session restore ───────────────────────────────────────────────────
    def show(self):
        self._mode_var.set(DB.ext_mode)
        self._rebuild_right_cols()
        if DB.ext_rows:
            self._populate_left()
            self._lbl_file.config(text=DB.ext_filename or "file loaded", fg=FG_R)
        if DB.ext_compilation:
            self._refresh_right()

    # ── Mode ──────────────────────────────────────────────────────────────
    def _mode_changed(self):
        DB.ext_mode = self._mode_var.get()
        DB.ext_compilation = []
        self._right_tree.delete(*self._right_tree.get_children())
        self._lbl_right_count.config(text="Rows: 0")
        self._btn_export.config(state="disabled")
        self._form_host.detach()
        self._hide_form_pane()
        self._rebuild_right_cols()

    def _rebuild_right_cols(self):
        if self._mode_var.get() == "Payment Reconciliation":
            cols   = self.RECON_COLS
            widths = [36, 90, 130, 100, 90, 110, 72]
        else:
            cols   = self.FAILED_COLS
            widths = [36, 150, 72]
        self._right_tree["columns"] = cols
        for c, w in zip(cols, widths):
            self._right_tree.heading(c, text=c)
            self._right_tree.column(c, width=w, minwidth=28, anchor="w")

    # ── CSV load ──────────────────────────────────────────────────────────
    def _load_csv(self):
        path = filedialog.askopenfilename(
            title="Select Freshservice CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader         = csv.DictReader(f)
                DB.ext_headers = list(reader.fieldnames or [])
                DB.ext_rows    = list(reader)
        except Exception as e:
            messagebox.showerror("Load Error", str(e))
            return
        if not DB.ext_rows:
            messagebox.showwarning("Empty", "The CSV file appears to be empty.")
            return
        DB.ext_col_ticket  = _detect(DB.ext_headers, _TICKET_KEYS)
        DB.ext_col_subject = _detect(DB.ext_headers, _SUBJECT_KEYS)
        DB.ext_col_desc    = _detect(DB.ext_headers, _DESC_KEYS)
        DB.ext_filename    = os.path.basename(path)
        self._populate_left()
        self._lbl_file.config(text=DB.ext_filename, fg=FG_R)
        self._hide_desc_pane()

    @staticmethod
    def _get(row: dict, *keys) -> str:
        for k in keys:
            if k in row:
                return row[k]
        return ""

    def _populate_left(self):
        self._left_tree.delete(*self._left_tree.get_children())
        self._row_map = {}
        for i, row in enumerate(DB.ext_rows, 1):
            tid     = row.get(DB.ext_col_ticket,  "")
            subj    = row.get(DB.ext_col_subject, "")
            created = self._get(row, *_CREATED_KEYS)
            done    = tid in DB.ext_left_done_ids
            tag     = "done" if done else ("alt" if i % 2 == 0 else "")
            iid     = str(i)
            self._left_tree.insert("", "end", iid=iid,
                                   values=(i, tid, subj, created), tags=(tag,))
            self._row_map[iid] = row
        self._lbl_left_count.config(text=f"Rows: {len(DB.ext_rows)}")

    # ── Left interactions ─────────────────────────────────────────────────
    def _show_desc(self, iid: str):
        row  = self._row_map.get(iid, {})
        tid  = row.get(DB.ext_col_ticket,  "")
        subj = row.get(DB.ext_col_subject, "")
        desc = row.get(DB.ext_col_desc,    "")
        self._desc_pane.show(tid, subj, desc)
        self._show_desc_pane()

    def _on_left_double(self, event):
        iid = self._left_tree.identify_row(event.y)
        if iid:
            self._left_tree.selection_set(iid)
            self._show_desc(iid)

    def _on_left_rclick(self, event):
        iid = self._left_tree.identify_row(event.y)
        if not iid:
            return
        self._left_tree.selection_set(iid)
        row     = self._row_map.get(iid, {})
        tid     = row.get(DB.ext_col_ticket, "")
        is_done = tid in DB.ext_left_done_ids
        menu = tk.Menu(self, tearoff=0, bg=BG, fg=TEXT, font=FONT_BODY,
                       activebackground=ACCENT, activeforeground=WHITE)
        menu.add_command(label="👁  View Description",
                         command=lambda: self._show_desc(iid))
        menu.add_command(label="➕  Add Entry",
                         command=lambda: self._open_form_for(iid))
        menu.add_separator()
        if is_done:
            menu.add_command(label="↩  Unmark Done",
                             command=lambda t=tid, i=iid: self._set_left_done(t, i, False))
        else:
            menu.add_command(label="✅  Mark as Done  🟢",
                             command=lambda t=tid, i=iid: self._set_left_done(t, i, True))
        menu.post(event.x_root, event.y_root)

    def _set_left_done(self, tid: str, iid: str, done: bool):
        if done:
            DB.ext_left_done_ids.add(tid)
            self._left_tree.item(iid, tags=("done",))
        else:
            DB.ext_left_done_ids.discard(tid)
            idx = int(iid)
            tag = "alt" if idx % 2 == 0 else ""
            self._left_tree.item(iid, tags=(tag,))

    # ── Right form — opens in FormHost pane above the tree ───────────────
    def _open_form(self):
        sel = self._left_tree.selection()
        if not sel:
            messagebox.showwarning(
                "No Ticket Selected",
                "Select a ticket from the left panel first.")
            return
        self._open_form_for(sel[0])

    def _open_form_for(self, iid: str):
        row = self._row_map.get(iid, {})
        tid = row.get(DB.ext_col_ticket, "")
        if self._mode_var.get() == "Payment Reconciliation":
            self._recon_form.prime(tid)
            self._recon_form.show()          # attaches to FormHost
        else:
            self._failed_form.prime(tid)
            self._failed_form.show()
        self._show_form_pane()               # expand top pane

    def _save_entry(self, data: dict):
        tid      = data.get("ticket_id", "")
        existing = [e.get("ticket_id", "") for e in DB.ext_compilation]
        if tid in existing:
            if not messagebox.askyesno(
                    "Duplicate",
                    f"Ticket #{tid} is already in the list. Add anyway?"):
                return
        txn_id = data.get("transaction_id", "").strip()
        if txn_id:
            existing_txns = [e.get("transaction_id", "").strip()
                             for e in DB.ext_compilation]
            if txn_id in existing_txns:
                if not messagebox.askyesno(
                        "Duplicate Transaction ID",
                        f"Transaction ID \"{txn_id}\" already exists in the list.\n\nAdd anyway?"):
                    return
        data["status"] = "Pending"
        DB.ext_compilation.append(data)
        self._refresh_right()
        # Auto-mark the corresponding left-tree row as done
        if tid:
            DB.ext_left_done_ids.add(tid)
            for liid, row in self._row_map.items():
                if row.get(DB.ext_col_ticket, "") == tid:
                    self._left_tree.item(liid, tags=("done",))
                    break

    # ── Right tree ────────────────────────────────────────────────────────
    def _refresh_right(self):
        self._right_tree.delete(*self._right_tree.get_children())
        mode = self._mode_var.get()
        for i, entry in enumerate(DB.ext_compilation, 1):
            done = entry.get("status") == "Done"
            tag  = "done" if done else ("alt" if i % 2 == 0 else "")
            if mode == "Payment Reconciliation":
                vals = (i,
                        entry.get("amount",         ""),
                        entry.get("phone",          ""),
                        entry.get("payment_date",   ""),
                        entry.get("payment_mode",   ""),
                        entry.get("transaction_id", ""),
                        entry.get("ticket_id",      ""))
            else:
                vals = (i, entry.get("phone", ""), entry.get("ticket_id", ""))
            self._right_tree.insert("", "end", iid=str(i),
                                    values=vals, tags=(tag,))
        total = len(DB.ext_compilation)
        self._lbl_right_count.config(text=f"Rows: {total}")
        self._btn_export.config(state="normal" if total else "disabled")

    def _on_right_rclick(self, event):
        iid = self._right_tree.identify_row(event.y)
        if not iid:
            return
        self._right_tree.selection_set(iid)
        idx = int(iid) - 1
        if not (0 <= idx < len(DB.ext_compilation)):
            return
        entry  = DB.ext_compilation[idx]
        status = entry.get("status", "Pending")
        menu = tk.Menu(self, tearoff=0, bg=BG, fg=TEXT, font=FONT_BODY,
                       activebackground=ACCENT, activeforeground=WHITE)
        if status != "Done":
            menu.add_command(
                label="✅  Mark as Done  🟢",
                command=lambda i=idx, iid_=iid:
                    self._set_right_status(i, iid_, "Done"))
        else:
            menu.add_command(
                label="↩  Unmark Done",
                command=lambda i=idx, iid_=iid:
                    self._set_right_status(i, iid_, "Pending"))
        menu.add_separator()
        menu.add_command(label="🗑  Remove Row",
                         command=lambda: self._remove_row(idx))
        menu.post(event.x_root, event.y_root)

    def _set_right_status(self, idx: int, iid: str, status: str):
        if 0 <= idx < len(DB.ext_compilation):
            DB.ext_compilation[idx]["status"] = status
            self._refresh_right()
            tid = DB.ext_compilation[idx].get("ticket_id", "")
            for liid, row in self._row_map.items():
                if row.get(DB.ext_col_ticket, "") == tid:
                    if status == "Done":
                        DB.ext_left_done_ids.add(tid)
                        self._left_tree.item(liid, tags=("done",))
                    else:
                        DB.ext_left_done_ids.discard(tid)
                        lx  = int(liid)
                        tag = "alt" if lx % 2 == 0 else ""
                        self._left_tree.item(liid, tags=(tag,))
                    break

    def _remove_row(self, idx: int):
        if 0 <= idx < len(DB.ext_compilation):
            DB.ext_compilation.pop(idx)
            self._refresh_right()

    def _clear_compilation(self):
        if not DB.ext_compilation:
            return
        if messagebox.askyesno("Clear All",
                               "Remove all rows from the compilation list?"):
            DB.ext_compilation = []
            self._refresh_right()

    def _export_excel(self):
        if not DB.ext_compilation:
            messagebox.showwarning("Nothing to Export", "Compilation list is empty.")
            return
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is not installed.\n\nRun:  pip install openpyxl")
            return
        mode  = self._mode_var.get()
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        label = "missing_payments" if mode == "Payment Reconciliation" else "failed_loans"
        name  = f"compiled_{label}_{ts}.xlsx"
        path  = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=name,
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = mode[:31]
            hdr_fill  = PatternFill("solid", fgColor="E8D5B0")
            done_fill = PatternFill("solid", fgColor="C8F2C2")
            hdr_font  = Font(bold=True, color="3A2F24")
            if mode == "Payment Reconciliation":
                headers = ["#", "Amount (UGX)", "Number Used to Pay",
                           "Payment Date", "Payment Mode", "Transaction ID", "Ticket #"]
            else:
                headers = ["#", "Phone Number", "Ticket #"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
            for i, entry in enumerate(DB.ext_compilation, 1):
                if mode == "Payment Reconciliation":
                    row_data = [i, entry.get("amount", ""), entry.get("phone", ""),
                                entry.get("payment_date", ""), entry.get("payment_mode", ""),
                                entry.get("transaction_id", ""), entry.get("ticket_id", "")]
                else:
                    row_data = [i, entry.get("phone", ""), entry.get("ticket_id", "")]
                ws.append(row_data)
                if entry.get("status") == "Done":
                    for cell in ws[ws.max_row]:
                        cell.fill = done_fill
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PANEL
# ─────────────────────────────────────────────────────────────────────────────
class TicketExtractorPanel(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._tab_panels: dict[str, tk.Frame] = {}
        self._active: str = ""
        try:
            _apply_ttk_styles(parent.winfo_toplevel())
        except Exception:
            pass
        self._build()

    def _build(self):
        top_bar = tk.Frame(self, bg="#e8d5b0", height=40,
                           highlightbackground=BORDER, highlightthickness=1)
        top_bar.pack(fill="x")
        top_bar.pack_propagate(False)

        tk.Label(top_bar, text="4G Workbench  ·  Ticket Extractor",
                 font=FONT_BOLD, bg="#e8d5b0", fg=TEXT).pack(
                     side="left", padx=10, pady=6)

        module_wrap = tk.Frame(top_bar, bg="#e8d5b0")
        module_wrap.pack(side="left", padx=(20, 0), pady=5)
        tk.Label(module_wrap, text="Module:", font=FONT_SM_B,
                 bg="#e8d5b0", fg=TEXT).pack(side="left", padx=(0, 6))

        _MODULE_NAMES = ["📋  Ticket Extraction", "🎫  Ticket Distributor"]
        self._module_var = tk.StringVar(value=_MODULE_NAMES[0])
        module_menu = _styled_combobox(
            module_wrap, self._module_var, _MODULE_NAMES, width=22, font=FONT_BODY)
        module_menu.pack(side="left")
        module_menu.bind("<<ComboboxSelected>>",
                         lambda _: self._switch(self._module_var.get()))

        content = tk.Frame(self, bg=BG)
        content.pack(fill="both", expand=True)

        panels: dict[str, tk.Frame] = {
            "🎫  Ticket Distributor": TicketsDistributionPanel(content),
            "📋  Ticket Extraction":  ExtractionPanel(content),
        }
        for name, panel in panels.items():
            self._tab_panels[name] = panel

        self._switch("📋  Ticket Extraction")

    def _switch(self, name: str):
        if self._active:
            self._tab_panels[self._active].pack_forget()
        self._tab_panels[name].pack(fill="both", expand=True)
        self._active = name
        self._module_var.set(name)
        if hasattr(self._tab_panels[name], "show"):
            self._tab_panels[name].show()

    def show(self):
        for panel in self._tab_panels.values():
            if hasattr(panel, "show"):
                panel.show()


# ─────────────────────────────────────────────────────────────────────────────
#  Standalone launcher
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    root.title("4G Workbench — Ticket Extractor")
    root.geometry("1200x750")
    root.configure(bg=BG)
    _apply_ttk_styles(root)
    panel = TicketExtractorPanel(root)
    panel.pack(fill="both", expand=True)
    root.mainloop()